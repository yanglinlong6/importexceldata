import pymysql
from openpyxl import load_workbook

if __name__ == '__main__':
    # 建立空字典
    result = {}
    workbook = load_workbook('E:\\data\\广州已安装GPS车辆.xlsx')
    sheets = workbook.get_sheet_names()
    ws = workbook.get_sheet_by_name(sheets[0])

    # 获取表格所有行和列，两者都是可迭代的
    rows = ws.rows
    columns = ws.columns

    conn = pymysql.connect(host='192.168.3.222',
                           port=3307,
                           user='root',
                           password='Msd^*$@online',
                           database='newgps',
                           charset='utf8')

    cursor = conn.cursor()

    # 清空对应关系表 重新导入
    sql00 = '''
    TRUNCATE TABLE d_vehicle_type_standno;
    '''
    cursor.execute(sql00)

    num = 0
    print('总行数:%s' % (ws.max_row + 1))
    for i in range(1, ws.max_row + 1):
        num = num + 1
        if i == 1:
            continue

        str03 = ws.cell(row=i, column=3).value
        str04 = ws.cell(row=i, column=4).value
        print('车架号VIN码:%s' % str03)
        print('车型:%s' % str04)
        if str03 is not None:
            sql = """
            INSERT INTO `newgps`.`d_vehicle_type_standno`(`standno`, `vehicle_type_name`, `create_time`) VALUES (%s, %s, now());
            """
            cursor.execute(sql, [str03, str04])
            data02 = cursor.fetchone()

        # print("Database version : %s " % data02)

    print("num : %s" % num)
    # 提交事务
    conn.commit()
    # 关闭光标对象
    cursor.close()

    # 关闭数据库连接
    conn.close()
