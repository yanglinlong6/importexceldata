from openpyxl import load_workbook

if __name__ == '__main__':
    # 建立空字典
    result = {}
    # uqueList = []
    workbook = load_workbook('E:\\data\\dianmian.xlsx')
    sheets = workbook.get_sheet_names()
    ws = workbook.get_sheet_by_name(sheets[0])

    # 获取表格所有行和列，两者都是可迭代的
    rows = ws.rows
    columns = ws.columns

    num = 0
    for i in range(1, ws.max_row + 1):
        num = num + 1
        if i == 1:
            continue
        sheet_one_order_one = ws.cell(row=i, column=1).value
        # sheet_one_order_two = ws.cell(row=i, column=2).value
        sheet_one_order_four = ws.cell(row=i, column=4).value
        print(sheet_one_order_one)
        # print(sheet_one_order_two)
        print(sheet_one_order_four)
        uqueDianMian = ''
        if isinstance(sheet_one_order_four, int):
            pass
        else:
            if sheet_one_order_four is not None:
                if ('</br>' in sheet_one_order_four):
                    # uqueDianMian = sheet_one_order_four.split("</br>")[1].split(":")[1]
                    strList = sheet_one_order_four.split("</br>")
                    for str in strList:
                        if ('店面名称' in str):
                            startindex = str.find('店面名称')
                            strSplit = str[startindex:]
                            print(strSplit)
                            if ':' in strSplit:
                                uqueDianMian = strSplit.split(":")[1]
                            elif '：' in strSplit:
                                uqueDianMian = strSplit.split("：")[1]
                elif ('<br/>' in sheet_one_order_four):
                    # uqueDianMian = sheet_one_order_four.split("</br>")[1].split(":")[1]
                    strList = sheet_one_order_four.split("<br/>")
                    for str in strList:
                        if ('店面名称' in str):
                            startindex = str.find('店面名称')
                            strSplit = str[startindex:]
                            print(strSplit)
                            if ':' in strSplit:
                                uqueDianMian = strSplit.split(":")[1]
                            elif '：' in strSplit:
                                uqueDianMian = strSplit.split("：")[1]
                elif ('店面名称' in sheet_one_order_four):
                    strList = sheet_one_order_four.splitlines()
                    for str in strList:
                        if ('店面名称' in str):
                            startindex = str.find('店面名称')
                            strSplit = str[startindex:]
                            print(strSplit)
                            if ':' in strSplit:
                                # print(strSplit.split(":"))
                                # uqueDianMian0 = strSplit.split(":")[0]
                                # print(uqueDianMian0)
                                uqueDianMian = strSplit.split(":")[1]
                            elif '—' in strSplit:
                                uqueDianMian = strSplit.split("—")[1]
                            elif '：' in strSplit:
                                uqueDianMian = strSplit.split("：")[1]

                uqueList = result.get(sheet_one_order_one)

                if uqueList is None:
                    uqueList = []
                    uqueList.append(uqueDianMian)
                    result[sheet_one_order_one] = uqueList
                    print('uqueList===== %s' % uqueList)
                else:
                    if (uqueDianMian not in uqueList):
                        uqueList.append(uqueDianMian)
                        result[sheet_one_order_one] = uqueList
                        print('uqueList===== %s' % uqueList)

    print('result===== %s' % result)
    print("num : %s" % num)
    for key, valueList in result.items():
        for value in valueList:
            print(key, value)
