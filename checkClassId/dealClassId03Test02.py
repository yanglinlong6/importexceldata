import pymysql
from openpyxl import load_workbook

conn = pymysql.connect(host='192.168.5.22',
                       port=3307,
                       user='user_yangll',
                       password='vGxw9jWg',
                       database='gps',
                       charset='utf8')
cursor = conn.cursor()

if __name__ == '__main__':
    # 建立空字典
    result = {}
    # workbook = load_workbook('E:\\data\\17-18年.xlsx')
    workbook = load_workbook('D:\\data01\\车辆档案列表 (导出).xlsx')
    sheets = workbook.get_sheet_names()
    ws = workbook.get_sheet_by_name(sheets[4])

    # 获取表格所有行和列，两者都是可迭代的
    rows = ws.rows
    columns = ws.columns

    snList = []
    num = 0
    print('总行数:%s' % (ws.max_row + 1))
    for i in range(1, ws.max_row + 1):
        num = num + 1
        print('====================')
        if i == 1:
            continue

        sqbh = ws.cell(row=i, column=1).value
        name = ws.cell(row=i, column=3).value
        standno = ws.cell(row=i, column=5).value
        sn = ws.cell(row=i, column=7).value
        groupName = ws.cell(row=i, column=9).value
        className = ws.cell(row=i, column=12).value
        print('工单号: %s' % sqbh)
        print('姓名:%s' % name)
        print('车架号:%s' % standno)
        print('设备号:%s' % sn)
        print('所属分组:%s' % groupName)
        print('所属分类:%s' % className)

        sql = '''
        SELECT dv.`standno` , dc.`className` ,dg.`groupName` ,dti.`sn` FROM d_vehicle dv 
        LEFT JOIN d_class dc ON dc.`classid` = dv.`classId` 
        LEFT JOIN d_group dg ON dg.`vehicleGroupId` = dc.`vehicleGroupId`
        LEFT JOIN d_track_info dti ON dti.`vehicleId` = dv.`vehicleId`
        WHERE dti.`sn` = %s;
        '''
        cursor.execute(sql, sn)
        resultData = cursor.fetchone()
        standno = resultData[0]
        classNameDB = resultData[1]
        groupNameDB = resultData[2]
        print('数据库分组:%s' % groupNameDB)
        print('数据库分类:%s' % classNameDB)
        sn = resultData[3]
        if classNameDB != className or groupNameDB != groupName:
            print('车架号===', standno, '设备号===', sn)
            snList.append(sn)

    print("num : %s" % num)
    print("snList : %s" % snList)
