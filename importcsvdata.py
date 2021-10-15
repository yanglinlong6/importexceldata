import csv
from datetime import datetime
from time import mktime

from openpyxl import load_workbook

if __name__ == '__main__':
    # 建立空字典
    result = {}
    workbook = load_workbook('E:\\data\\gps.xlsx')
    sheets = workbook.get_sheet_names()
    ws = workbook.get_sheet_by_name(sheets[0])

    # 获取表格所有行和列，两者都是可迭代的
    rows = ws.rows
    columns = ws.columns

    for i in range(1, ws.max_row + 1):
        sheet_one_order_one = ws.cell(row=i, column=1).value
        sheet_one_order_two = ws.cell(row=i, column=2).value
        print(sheet_one_order_one)
        print(sheet_one_order_two)
        result[sheet_one_order_two] = sheet_one_order_one
        print("get=====", result.get(sheet_one_order_two))

    # 读取csv至字典
    csvFile = open('E:\\data\\20210118-002.csv', "r", encoding="utf8")
    reader = csv.reader(csvFile)

    sql = 'INSERT INTO `data_obd_v3`.`glsx_journey` (`user_id`, `begin_gps_lat`, `begin_gps_lon`, `begin_address`, `end_gps_lat`, `end_gps_lon`, `end_address`, `begin_time`, `create_time`, `end_time`, `mileage`,  `sn`, `total_time`, `del_status`, `gps_begin_time`, `gps_end_time`, `gps_total_time`, `imei`) VALUES '
    insertStr = ''
    num = 0
    snList = ['1790701405', '1701027410', '1701218317', '1701107589', '1790701405', '1701027410', '1700923487',
              '1701014726', '1790325630', '1780904583']

    count = 0
    uqueList = []

    for item in reader:
        count = count + 1

        # 忽略第一行
        if reader.line_num == 1:
            continue

        # print(item[0])
        # print(item[1])
        # print(item[2])
        # print(item[3])
        userId = result.get(item[0])
        print('sn=====', item[0])
        print('~~~~~~~~~~~~~~~~~')
        print('userId=====', userId)
        print('~~~~~~~~~~~~~~~~~')

        if (item[0] not in snList and userId != None and len(userId) != 0):
            uqueKey = userId + item[2]
            if (uqueKey not in uqueList):
                d1 = datetime.strptime(item[1], "%Y-%m-%d %H:%M:%S")
                timeStamp1 = int(mktime(d1.timetuple()))

                d2 = datetime.strptime(item[2], "%Y-%m-%d %H:%M:%S")
                timeStamp2 = int(mktime(d2.timetuple()))

                str_list = [userId, item[30].split(',')[1], item[30].split(',')[0], item[29], item[32].split(',')[1],
                            item[32].split(',')[0], item[31], item[1], item[2], item[2], item[11], item[0],
                            str(timeStamp2 - timeStamp1), str(0), item[1], item[2], str(timeStamp2 - timeStamp1),
                            item[0]]

                # insertStr += '(' + userId + ', ' + item[30].split(',')[1] + ',' + item[30].split(',')[0] + ',' + item[
                #     29] + ', ' + \
                #              item[32].split(',')[1] + ',' + item[32].split(',')[0] + ',' + item[31] + ',' + item[1] + ',' + \
                #              item[2] + ',' + item[2] + ', ' + item[11] + ',' + item[0] + ', ' + str(
                #     timeStamp2 - timeStamp1) + ', 0, ' + item[1] + ',' + item[2] + ', ' + str(
                #     timeStamp2 - timeStamp1) + ', ' + \
                #              item[0] + '),'
                uqueList.append(userId + item[2])

                join_list = "','".join(str_list)
                insertStr += "('" + join_list + "'),"
                num = num + 1
                while (num == 100000):
                    sql_insert_str = sql + insertStr
                    print(sql_insert_str[0:len(sql_insert_str) - 1] + ';')
                    with open("E:\\data\\runSql01.sql", "a") as sob:
                        sob.write(sql_insert_str[0:len(sql_insert_str) - 1] + ';\n')
                    insertStr = ''
                    num = 0

    sql_insert_str = sql + insertStr
    print(sql_insert_str[0:len(sql_insert_str) - 1] + ';')
    with open("E:\\data\\runSql01.sql", "a") as sob:
        sob.write(sql_insert_str[0:len(sql_insert_str) - 1] + ';\n')

    print(count)
    csvFile.close()

# sn 0 imei
# bgtime 1 begin_time
# edtime 2 end_time
# dristarth 3
# driendh 4
# dristartm 5
# driendm 6
# joursign 7
# isjourbg 8
# jourmil 9
# jourdriti 10
# mileage 11
# dritime 12
# drispeed 13
# hispeedct 14
# hispeedti 15
# mbusyti 16
# mbusymil 17
# nbusyti 18
# nbusymil 19
# upnightmil 20
# upnightti 21
# downnightmil 22
# downnightti 23
# iswkday 24
# isbadwth0 25
# isbadwth1 26
# isbadwth2 27
# curvratio 28
# startcity 29
# startlnglat 30
# stopcity 31
# stoplnglat 32
# iscrosscity 33
# p0ti 34
# p1ti 35
# p2ti 36
# p3ti 37
# p4ti 38
# p5ti 39
# p6ti 40
# p7ti 41
# p8ti 42
# p9ti 43
# p10ti 44
# p11ti 45
# p12ti 46
# p13ti 47
# p14ti 48
# p15ti 49
# p16ti 50
# p17ti 51
# p18ti 52
# p19ti 53
# p20ti 54
# p21ti 55
# p22ti 56
# p23ti 57
# month 58
# day 59
#
