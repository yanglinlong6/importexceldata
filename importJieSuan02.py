import pymysql
from openpyxl import load_workbook

if __name__ == '__main__':
    # 建立空字典
    result = {}
    # workbook = load_workbook('E:\\data\\17-18年.xlsx')
    workbook = load_workbook('E:\\data\\19-20年.xlsx')
    sheets = workbook.get_sheet_names()
    ws = workbook.get_sheet_by_name(sheets[0])

    # 获取表格所有行和列，两者都是可迭代的
    rows = ws.rows
    columns = ws.columns

    conn = pymysql.connect(host='192.168.3.222',
                           port=3307,
                           user='root',
                           password='Msd^*$@online',
                           database='newgps',
                           charset='utf8')

    cursor = conn.cursor()
    num = 0
    print('总行数:%s' % (ws.max_row + 1))
    for i in range(1, ws.max_row + 1):
        num = num + 1
        if i == 1:
            continue

        str01 = ws.cell(row=i, column=3).value
        str02 = ws.cell(row=i, column=5).value
        str03 = ws.cell(row=i, column=9).value
        str04 = ws.cell(row=i, column=10).value
        str05 = ws.cell(row=i, column=18).value
        str06 = ws.cell(row=i, column=19).value
        print('订单日期: %s' % str01)
        print('客户名称:%s' % str02)
        print('产品名称:%s' % str03)
        print('数量:%s' % str04)
        print('单价:%s' % str05)
        print('销售总额:%s' % str06)

        sql = 'INSERT INTO `newgps`.`d_settlement`(`date`, `customer_information`, `product_name`, `equipment_num`, `equipment_cost`, `service_charge`, `unit_price`, `total_amount`) VALUES (%s, %s, %s, %s, %s, %s, %s, %s);'

        cursor.execute(sql, [str01, str02, str03, str04, 0, 0, str05, str06])
        data02 = cursor.fetchone()
        print("Database version : %s " % data02)

    print("num : %s" % num)
    # 提交事务
    conn.commit()
    # 关闭光标对象
    cursor.close()

    # 关闭数据库连接
    conn.close()
