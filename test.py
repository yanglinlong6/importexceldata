from openpyxl import load_workbook

if __name__ == '__main__':
    # 建立空字典
    result = {}
    workbook  = load_workbook('E:\\data\\gps.xlsx')
    sheets = workbook.get_sheet_names()
    ws  = workbook.get_sheet_by_name(sheets[0])

    # 获取表格所有行和列，两者都是可迭代的
    rows = ws.rows
    columns = ws.columns

    for i in range(1,ws.max_row+1):
        sheet_one_order_one = ws.cell(row=i, column=1).value
        sheet_one_order_two = ws.cell(row=i, column=2).value
        # print(sheet_one_order_one)
        # print(sheet_one_order_two)
        result[sheet_one_order_one]=sheet_one_order_two
        print(result.get(sheet_one_order_one))
